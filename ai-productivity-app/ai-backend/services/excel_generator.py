import os
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import json

logger = logging.getLogger(__name__)

class ExcelGenerator:
    def __init__(self):
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def generate_plan_excel(self, plan_data):
        """Generate Excel file for event plan."""
        try:
            # Create exports directory if it doesn't exist
            os.makedirs('exports', exist_ok=True)
            
            filename = f"event_plan_{plan_data.get('plan_id', 'unknown')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join('exports', filename)
            
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Event Details Sheet
            self._create_event_details_sheet(wb, plan_data)
            
            # Timeline Sheet
            self._create_timeline_sheet(wb, plan_data)
            
            # Roles Sheet
            self._create_roles_sheet(wb, plan_data)
            
            # Budget Sheet
            self._create_budget_sheet(wb, plan_data)
            
            # Tasks Sheet
            self._create_tasks_sheet(wb, plan_data)
            
            # Save workbook
            wb.save(filepath)
            
            logger.info(f"Excel file generated successfully: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Failed to generate Excel file: {e}")
            raise e
    
    def _create_event_details_sheet(self, wb, plan_data):
        """Create event details sheet."""
        ws = wb.create_sheet("Event Details")
        
        # Headers
        ws['A1'] = "Event Management Plan"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:B1')
        
        # Event details
        event_details = plan_data.get('event_details', {})
        details = [
            ['Event Type', plan_data.get('event_type', 'General').title()],
            ['Date', event_details.get('date', 'TBD')],
            ['Venue', event_details.get('venue', 'TBD')],
            ['Expected Attendees', str(event_details.get('attendees', 'TBD'))],
            ['Duration', event_details.get('duration', 'TBD')],
            ['Budget', event_details.get('budget', 'TBD')],
            ['Generated On', datetime.now().strftime('%B %d, %Y at %I:%M %p')]
        ]
        
        row = 3
        for detail in details:
            ws[f'A{row}'] = detail[0]
            ws[f'B{row}'] = detail[1]
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
    
    def _create_timeline_sheet(self, wb, plan_data):
        """Create timeline sheet."""
        ws = wb.create_sheet("Timeline")
        
        # Headers
        headers = ['Phase', 'Duration', 'Tasks', 'Dependencies']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        # Timeline data
        ai_plan = plan_data.get('ai_plan', {})
        timeline = ai_plan.get('timeline', [])
        
        row = 2
        if isinstance(timeline, list):
            for phase_data in timeline:
                if isinstance(phase_data, dict):
                    phase = phase_data.get('phase', 'Phase')
                    duration = phase_data.get('duration', 'TBD')
                    tasks = phase_data.get('tasks', [])
                    dependencies = phase_data.get('dependencies', [])
                    
                    ws.cell(row=row, column=1, value=phase).border = self.border
                    ws.cell(row=row, column=2, value=duration).border = self.border
                    ws.cell(row=row, column=3, value='; '.join(tasks) if tasks else 'No tasks').border = self.border
                    ws.cell(row=row, column=4, value='; '.join(dependencies) if dependencies else 'None').border = self.border
                    
                    row += 1
        
        # Auto-adjust column widths
        for col in range(1, 5):
            ws.column_dimensions[chr(64 + col)].width = 25
    
    def _create_roles_sheet(self, wb, plan_data):
        """Create roles sheet."""
        ws = wb.create_sheet("Team Roles")
        
        # Headers
        headers = ['Role Title', 'Responsibilities', 'Required Skills', 'Priority']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        # Roles data
        ai_plan = plan_data.get('ai_plan', {})
        roles = ai_plan.get('roles', [])
        
        row = 2
        if isinstance(roles, list):
            for role_data in roles:
                if isinstance(role_data, dict):
                    title = role_data.get('title', 'Role')
                    responsibilities = role_data.get('responsibilities', [])
                    skills = role_data.get('skills', [])
                    priority = role_data.get('priority', 'Medium')
                    
                    ws.cell(row=row, column=1, value=title).border = self.border
                    ws.cell(row=row, column=2, value='; '.join(responsibilities) if responsibilities else 'TBD').border = self.border
                    ws.cell(row=row, column=3, value='; '.join(skills) if skills else 'TBD').border = self.border
                    ws.cell(row=row, column=4, value=priority).border = self.border
                    
                    row += 1
        
        # Auto-adjust column widths
        for col in range(1, 5):
            ws.column_dimensions[chr(64 + col)].width = 30
    
    def _create_budget_sheet(self, wb, plan_data):
        """Create budget sheet."""
        ws = wb.create_sheet("Budget")
        
        # Headers
        headers = ['Category', 'Amount', 'Percentage', 'Notes']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        # Budget data
        ai_plan = plan_data.get('ai_plan', {})
        budget = ai_plan.get('budget_breakdown', {})
        
        row = 2
        total_amount = 0
        
        if isinstance(budget, dict) and 'categories' in budget:
            for category_data in budget['categories']:
                if isinstance(category_data, dict):
                    category = category_data.get('category', 'Category')
                    amount = category_data.get('amount', '₹0')
                    percentage = category_data.get('percentage', 0)
                    notes = category_data.get('notes', '')
                    
                    ws.cell(row=row, column=1, value=category).border = self.border
                    ws.cell(row=row, column=2, value=amount).border = self.border
                    ws.cell(row=row, column=3, value=f"{percentage}%").border = self.border
                    ws.cell(row=row, column=4, value=notes).border = self.border
                    
                    row += 1
        
        # Auto-adjust column widths
        for col in range(1, 5):
            ws.column_dimensions[chr(64 + col)].width = 20
    
    def _create_tasks_sheet(self, wb, plan_data):
        """Create tasks sheet."""
        ws = wb.create_sheet("Task Checklist")
        
        # Headers
        headers = ['Task', 'Assigned Role', 'Priority', 'Deadline', 'Status', 'Notes']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        # Extract tasks from timeline and roles
        ai_plan = plan_data.get('ai_plan', {})
        timeline = ai_plan.get('timeline', [])
        
        row = 2
        if isinstance(timeline, list):
            for phase_data in timeline:
                if isinstance(phase_data, dict):
                    phase = phase_data.get('phase', 'Phase')
                    tasks = phase_data.get('tasks', [])
                    
                    for task in tasks:
                        ws.cell(row=row, column=1, value=task).border = self.border
                        ws.cell(row=row, column=2, value='TBD').border = self.border
                        ws.cell(row=row, column=3, value='Medium').border = self.border
                        ws.cell(row=row, column=4, value='TBD').border = self.border
                        ws.cell(row=row, column=5, value='Pending').border = self.border
                        ws.cell(row=row, column=6, value=f'Part of {phase}').border = self.border
                        
                        row += 1
        
        # Auto-adjust column widths
        for col in range(1, 7):
            ws.column_dimensions[chr(64 + col)].width = 25
